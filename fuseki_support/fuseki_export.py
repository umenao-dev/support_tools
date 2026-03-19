import argparse
import csv
import datetime as dt
import json
import os
import re
import sys
from typing import Dict, Iterable, List, Optional, Tuple

import requests
from openpyxl import Workbook

EXCEL_MAX_ROWS = 1_048_576


def _normalize_dataset_name(name: str) -> str:
    return name.lstrip("/")


def _datasets_from_json(data: dict) -> List[str]:
    datasets = []
    items = None
    if isinstance(data, dict):
        items = data.get("datasets") or data.get("dataset") or data.get("Dataset")
    if items is None:
        return datasets
    for item in items:
        if isinstance(item, dict):
            name = (
                item.get("ds.name")
                or item.get("name")
                or item.get("dsName")
                or item.get("dataset")
            )
            if isinstance(name, str):
                datasets.append(_normalize_dataset_name(name))
        elif isinstance(item, str):
            datasets.append(_normalize_dataset_name(item))
    return [d for d in datasets if d]


def get_datasets(base_url: str, timeout: int) -> List[str]:
    url = base_url.rstrip("/") + "/$/datasets"
    resp = requests.get(url, headers={"Accept": "application/json"}, timeout=timeout)
    resp.raise_for_status()
    try:
        data = resp.json()
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Dataset list response is not JSON: {url}") from exc
    datasets = _datasets_from_json(data)
    if not datasets:
        raise RuntimeError("No datasets found in /$/datasets response.")
    return datasets


def sparql_query(endpoint: str, query: str, timeout: int) -> dict:
    resp = requests.post(
        endpoint,
        data={"query": query},
        headers={"Accept": "application/sparql-results+json"},
        timeout=timeout,
    )
    resp.raise_for_status()
    return resp.json()


def list_graphs(endpoint: str, timeout: int) -> Tuple[bool, List[str]]:
    ask_default = "ASK { ?s ?p ?o }"
    ask_result = sparql_query(endpoint, ask_default, timeout)
    default_has_triples = bool(ask_result.get("boolean"))

    query = "SELECT DISTINCT ?g WHERE { GRAPH ?g { ?s ?p ?o } } ORDER BY ?g"
    data = sparql_query(endpoint, query, timeout)
    graphs = []
    for binding in data.get("results", {}).get("bindings", []):
        g = binding.get("g", {})
        if "value" in g:
            graphs.append(g["value"])
    return default_has_triples, graphs


def list_recent_graphs(endpoint: str, limit: int, timeout: int) -> List[str]:
    query = (
        "SELECT DISTINCT ?g WHERE { "
        "GRAPH ?g { ?s ?p ?o } "
        "FILTER(REGEX(STR(?g), \"/\\\\d{14}/?$\")) "
        "} ORDER BY DESC(STR(?g)) "
        f"LIMIT {limit}"
    )
    data = sparql_query(endpoint, query, timeout)
    graphs = []
    for binding in data.get("results", {}).get("bindings", []):
        g = binding.get("g", {})
        if "value" in g:
            graphs.append(g["value"])
    return graphs


def _extract_timestamp_id(graph_uri: str) -> Optional[str]:
    match = re.search(r"/(\d{14})/?$", graph_uri)
    if not match:
        return None
    return match.group(1)


def _term_fields(binding: Dict[str, str]) -> Tuple[str, str, str, str]:
    value = binding.get("value", "")
    term_type = binding.get("type", "")
    datatype = binding.get("datatype", "")
    lang = binding.get("xml:lang", "")
    return value, term_type, datatype, lang


def iter_triples(endpoint: str, graph: Optional[str], page_size: int, timeout: int):
    offset = 0
    while True:
        if graph is None:
            query = (
                "SELECT ?s ?p ?o WHERE { ?s ?p ?o } "
                f"ORDER BY ?s ?p ?o LIMIT {page_size} OFFSET {offset}"
            )
        else:
            query = (
                "SELECT ?s ?p ?o WHERE { GRAPH <" + graph + "> { ?s ?p ?o } } "
                f"ORDER BY ?s ?p ?o LIMIT {page_size} OFFSET {offset}"
            )
        data = sparql_query(endpoint, query, timeout)
        bindings = data.get("results", {}).get("bindings", [])
        if not bindings:
            break
        for row in bindings:
            yield row
        offset += page_size


def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Export all graphs from all Fuseki datasets into CSV and Excel.")
    parser.add_argument(
        "--base-url", default="http://127.0.0.1:3030", help="Fuseki base URL"
    )
    parser.add_argument(
        "--datasets",
        help="Comma-separated dataset names to export (if omitted, export all)",
    )
    parser.add_argument(
        "--dataset",
        action="append",
        help="Dataset name to export (repeatable).",
    )
    parser.add_argument("--out-dir", default="out", help="Output directory")
    parser.add_argument("--page-size", type=int, default=10000, help="SPARQL page size")
    parser.add_argument("--timeout", type=int, default=60, help="HTTP timeout (s)")
    parser.add_argument(
        "--recent-limit",
        type=int,
        default=5,
        help="Show N recent graph URIs (by timestamp suffix) before export",
    )
    parser.add_argument(
        "--recent-only",
        action="store_true",
        help="Only show recent graph URIs and skip CSV/Excel export",
    )
    args = parser.parse_args()

    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    ensure_dir(args.out_dir)
    csv_path = os.path.join(args.out_dir, f"fuseki_export_{timestamp}.csv")
    xlsx_path = os.path.join(args.out_dir, f"fuseki_export_{timestamp}.xlsx")

    selected = []
    if args.datasets:
        selected.extend([d.strip() for d in args.datasets.split(",") if d.strip()])
    if args.dataset:
        selected.extend([d for d in args.dataset if d])

    if selected:
        datasets = [_normalize_dataset_name(d) for d in selected]
    else:
        datasets = get_datasets(args.base_url, args.timeout)

    recent_entries = []
    if args.recent_limit and args.recent_limit > 0:
        for dataset in datasets:
            endpoint = args.base_url.rstrip("/") + "/" + dataset + "/sparql"
            graphs = list_recent_graphs(endpoint, args.recent_limit, args.timeout)
            for graph in graphs:
                ts = _extract_timestamp_id(graph)
                if ts:
                    recent_entries.append((ts, dataset, graph))

        recent_entries.sort(key=lambda x: x[0], reverse=True)
        if recent_entries:
            print(f"Recent graphs (top {args.recent_limit})")
            for ts, dataset, graph in recent_entries[: args.recent_limit]:
                print(f"  {ts} | {dataset} | {graph}")
        else:
            print(
                f"Recent graphs (top {args.recent_limit}) - none matched timestamp URI pattern."
            )

    if args.recent_only:
        return 0

    header = [
        "dataset",
        "graph",
        "subject",
        "predicate",
        "object",
        "subject_type",
        "predicate_type",
        "object_type",
        "object_datatype",
        "object_lang",
    ]

    wb = Workbook(write_only=True)
    sheet_index = 1
    ws = wb.create_sheet(title=f"triples_{sheet_index}")
    ws.append(header)
    excel_row_count = 1

    def excel_append(row: List[str]) -> None:
        nonlocal ws, sheet_index, excel_row_count
        if excel_row_count >= EXCEL_MAX_ROWS:
            sheet_index += 1
            ws = wb.create_sheet(title=f"triples_{sheet_index}")
            ws.append(header)
            excel_row_count = 1
        ws.append(row)
        excel_row_count += 1

    total_rows = 0

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)

        for dataset in datasets:
            endpoint = args.base_url.rstrip("/") + "/" + dataset + "/sparql"
            print(f"Dataset: {dataset} -> {endpoint}")

            default_has_triples, graphs = list_graphs(endpoint, args.timeout)
            graph_list = []
            if default_has_triples:
                graph_list.append(None)
            graph_list.extend(graphs)

            for graph in graph_list:
                graph_label = "DEFAULT" if graph is None else graph
                print(f"  Graph: {graph_label}")

                for binding in iter_triples(endpoint, graph, args.page_size, args.timeout):
                    s_val, s_type, _, _ = _term_fields(binding.get("s", {}))
                    p_val, p_type, _, _ = _term_fields(binding.get("p", {}))
                    o_val, o_type, o_dt, o_lang = _term_fields(binding.get("o", {}))

                    row = [
                        dataset,
                        graph_label,
                        s_val,
                        p_val,
                        o_val,
                        s_type,
                        p_type,
                        o_type,
                        o_dt,
                        o_lang,
                    ]
                    writer.writerow(row)
                    excel_append(row)
                    total_rows += 1

    wb.save(xlsx_path)

    print(f"Done. rows={total_rows}")
    print(f"CSV : {csv_path}")
    print(f"XLSX: {xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
